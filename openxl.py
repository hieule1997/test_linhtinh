import openpyxl

theFile = openpyxl.load_workbook('example3.xlsx')
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))


for x in allSheetNames:
    print("Current sheet name is {}" .format(x))
    currentSheet = theFile[x]
    print(currentSheet['B4'].value)